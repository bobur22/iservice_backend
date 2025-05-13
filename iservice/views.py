from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from user.models import User
from user.forms import UserForm
from .models import Customer, Task, Location
from .forms import CustomerForm, TasksForm
from .filters import CustomerFilter, UserFilter, TasksFilter
import openpyxl
from decimal import Decimal
from django.http import HttpResponse


@login_required(login_url="login")
def dashboard_view(request):
    return render(request, 'dashboard.html')


@login_required(login_url="login")
def calendar_view(request):
    return render(request, 'calendar.html')


@login_required(login_url="login")
def employees_view(request):
    users = User.objects.all().filter(role='employee')
    users_filter = UserFilter(request.GET, queryset=users)
    u_count = users_filter.qs.count()
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('employees')
    else:
        form = UserForm()
    return render(request, 'employees.html', {
        'form': form,
        'u_count': u_count,
        'filter': users_filter,
    })


def employees_delete(request, id):
    users = User.objects.get(id=id)
    users.delete()
    return redirect('employees')


@login_required(login_url="login")
def employees_detail_view(request, id):
    user_d = User.objects.filter(role='employee').get(id=id)
    form = UserForm(instance=user_d)
    if request.method == "POST":
        form = UserForm(request.POST, instance=user_d)
        if form.is_valid():
            form.save()
            return redirect("employees")
    return render(request, 'employees_detail.html', {'form': form, })


@login_required(login_url="login")
def customers_view(request):
    if request.user.is_superuser:
        user_l = None
        customers = Customer.objects.all()
    else:
        user_l = User.objects.get(email=request.user.email).location
        customers = Customer.objects.all().filter(location=user_l)

    customers_filter = CustomerFilter(request.GET, queryset=customers)
    c_count = customers_filter.qs.count()

    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer_f = form.save(commit=False)
            customer_f.location = user_l
            customer_f.save()
            return redirect('customers')
    else:
        form = CustomerForm()
    return render(request, 'customers.html', {'form': form, 'c_count': c_count, 'filter': customers_filter})


def customer_delete(request, id):
    customer = Customer.objects.get(id=id)
    customer.delete()
    return redirect('customers')


@login_required(login_url="login")
def customers_detail_view(request, id):
    customers_d = Customer.objects.get(id=id)
    form = CustomerForm(instance=customers_d)

    if request.method == "POST":
        form = CustomerForm(request.POST, instance=customers_d)
        if form.is_valid():
            form.save()
            return redirect("customers")
    return render(request, 'customers_detail.html', {'form': form, })


@login_required(login_url="login")
def download_customers_excel(request):
    if request.user.is_superuser:
        customers = Customer.objects.all()
    else:
        user_l = request.user.location
        customers = Customer.objects.filter(location=user_l)

    # Apply filter
    customers_filter = CustomerFilter(request.GET, queryset=customers)
    filtered_customers = customers_filter.qs

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Write headers
    ws.append([
        "ID", "Full Name", "Phone", "Email", "Location", "Created At"
    ])

    n = 1
    # Write customer data
    for customer in filtered_customers:
        ws.append([
            n,
            customer.full_name,
            customer.p_number,
            customer.email,
            str(customer.location),  # convert FK to string
            customer.created_at.strftime("%Y-%m-%d %H:%M"),
        ])
        n += 1

    # Prepare HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=filtered_customers.xlsx'

    wb.save(response)
    return response


@login_required(login_url="login")
def tasks_view(request):
    user = request.user
    base_qs = Task.objects.all() if user.is_superuser else Task.objects.filter(location=user.location)

    tasks_filter = TasksFilter(request.GET, queryset=base_qs, request=request)
    filtered_qs = tasks_filter.qs

    total_tasks = filtered_qs.count()
    active_tasks = filtered_qs.filter(role__in=['To Do', 'In Progress', 'In Review']).count()

    if request.method == 'POST':
        form = TasksForm(request.POST, user=user)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = user
            task.location = user.location  # enforce user-location integrity
            task.save()
            return redirect('tasks')
    else:
        form = TasksForm(user=user)

    return render(request, 'tasks.html', {
        'filter': tasks_filter,
        'form': form,
        't_count': total_tasks,
        'ta_count': active_tasks,
    })


def task_delete(request, id):
    task = Task.objects.get(id=id)
    task.delete()
    return redirect('tasks')


@login_required(login_url="login")
def tasks_detail_view(request, id):
    task_d = Task.objects.get(id=id)
    if request.method == "POST":
        form = TasksForm(request.POST, instance=task_d)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            task.deadline = task_d.deadline
            task.save()
            return redirect('tasks')
    else:
        form = TasksForm(instance=task_d, user=request.user)
    return render(request, 'task_detail.html', {'form': form, "request": request,})


@login_required(login_url="login")
def tasks_archive_view(request):
    user = request.user
    base_qs = (
        Task.objects.filter(role='Done')
        if user.is_superuser else
        Task.objects.filter(location=user.location, role='Done')
    )
    tasks_filter = TasksFilter(request.GET, queryset=base_qs, request=request)
    total_tasks = tasks_filter.qs.count()

    ta_count = total_tasks - tasks_filter.qs.filter(payment_t='Process').count()

    return render(request, 'tasks_done.html', {
        'filter': tasks_filter,
        't_count': total_tasks,
        'ta_count': ta_count,
    })


@login_required(login_url="login")
def tasks_archive_download_excel(request):
    user = request.user

    base_qs = (
        Task.objects.filter(role='Done')
        if user.is_superuser else
        Task.objects.filter(location=user.location, role='Done')
    )

    tasks_filter = TasksFilter(request.GET, queryset=base_qs, request=request)
    filtered_tasks = tasks_filter.qs

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Done Tasks"

    headers = [
        'ID', 'Title', 'Price', 'Customer', 'Phone Number', 'Payment Type',
        'Cost', 'Program', 'Program Sum', 'Status', 'Created at', 'Deadline',
        'Location', 'Overall Sum'
    ]
    ws.append(headers)

    n = 1
    for task in filtered_tasks:
        ws.append([
            n,
            task.task_title,
            task.price,
            task.customer.full_name if task.customer else '',
            task.customer.p_number if task.customer else '',
            task.payment_t,
            task.cost,
            task.program,
            task.program_sum,
            task.role,
            task.created_at.strftime('%Y-%m-%d') if task.created_at else '',
            task.deadline.strftime('%Y-%m-%d') if task.deadline else '',
            task.location.location_name if task.location else '',
            task.price + task.cost + task.program_sum
        ])
        n += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=done_tasks.xlsx'
    wb.save(response)
    return response


@login_required(login_url="login")
def locations_view(request):
    tasks = Task.objects.all()

    tasks_filter = TasksFilter(request.GET, queryset=tasks, request=request)
    filtered_tasks = tasks_filter.qs
    t_count = tasks_filter.qs.count()

    location_id = request.GET.get('location')

    if location_id:
        try:
            location_obj = Location.objects.get(id=location_id)
            location_name = location_obj.location_name
        except Location.DoesNotExist:
            location_name = 'All'
    else:
        location_name = 'All'

    context = {
        't_count': t_count,
        'location_n': location_name,
        'filter': tasks_filter,
        'tasks_status_t': filtered_tasks.filter(role='To Do'),
        'tasks_status_p': filtered_tasks.filter(role='In Progress'),
        'tasks_status_r': filtered_tasks.filter(role='In Review'),
        'tasks_status_d': filtered_tasks.filter(role='Done'),
    }

    return render(request, 'locations.html', context)


@login_required(login_url="login")
def download_excel(request):
    tasks = Task.objects.all()
    tasks_filter = TasksFilter(request.GET, queryset=tasks, request=request)
    filtered_tasks = tasks_filter.qs

    # Create Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered Tasks"

    # Add header
    headers = ['ID', 'Title', 'Price', 'Customer', 'Phone Number', 'Payment Type', 'Cost', 'Program', 'Program Sum',
               'Status', 'Created at', 'Deadline', 'Location', 'Overall Sum']
    ws.append(headers)

    n = 1
    for task in filtered_tasks:
        ws.append([
            n,
            task.task_title,
            task.price,
            task.customer.full_name if task.customer else '',
            task.customer.p_number if task.customer else '',
            task.payment_t,
            task.cost,
            task.program,
            task.program_sum,
            task.role,
            task.created_at,
            task.deadline.strftime('%Y-%m-%d') if task.deadline else '',
            task.location.location_name if task.location else '',
            (task.price or Decimal('0')) + (task.cost or Decimal('0')) + (task.program_sum or Decimal('0'))
        ])
        n += 1

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=filtered_tasks.xlsx'
    wb.save(response)
    return response
